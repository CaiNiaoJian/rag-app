"""`.xls` 的纯 Python 兜底转换：xlrd 读 BIFF → openpyxl 写 xlsx（03 章 §4 备选路径）。

为什么存在：LibreOffice 随包（M2/D3，L 级工作）落地前，干净 Windows 机上
``find_soffice()`` 三路查找全部落空，导入 .xls 必抛 E03——七格式实际只有四格式。
xlrd（BSD-3，2.x 起仅支持 .xls）成本只有几十 KB，先让 .xls 真正有产出。

刻意选择「转成 xlsx 文件再走 xlsx_parser」而不是写第二个表格解析器：
解析层的约定是旧格式一律先归一化（parsers/__init__.py 约定 1），区域切分、
表头启发式、合并单元格、降级链、指标口径全部只维护一套。

能力边界（诚实声明，写进 convert_chain 的 ``(xlrd)`` 后缀即为溯源标记）：
- 保留：单元格值（文本/数值/日期/布尔/错误码）、合并区域、字体加粗
  （表头启发式依赖它）、sheet 名与顺序；
- 不保留：图片、图表、批注 —— xlrd 根本读不出这些，LibreOffice 路径可以。
  因此查找顺序仍是 soffice 优先，本模块只做兜底。
"""

from __future__ import annotations

from datetime import time
from pathlib import Path
from typing import Any

from docfactory.errors import DocFactoryError

# xlrd 的单元格类型常量（避免 import * 又不想逐处写 magic number）
_XL_TEXT = 1
_XL_NUMBER = 2
_XL_DATE = 3
_XL_BOOLEAN = 4
_XL_ERROR = 5


def convert_xls_to_xlsx(src: Path, out_dir: Path) -> Path:
    """把 .xls 转写为 ``out_dir/{stem}.xlsx``，返回产物路径。

    失败语义与其它导入路径对齐：密码保护 → E02，损坏/非 .xls → E01。
    ``out_dir`` 由调用方管理生命周期（parse_document 的 staging 临时目录）。
    """
    book, has_formatting = _open_xls(Path(src))
    try:
        workbook = _build_workbook(book, has_formatting)
    finally:
        book.release_resources()

    out_dir = Path(out_dir)
    out_dir.mkdir(parents=True, exist_ok=True)
    dest = out_dir / f"{Path(src).stem}.xlsx"
    workbook.save(dest)
    return dest


def _open_xls(src: Path) -> tuple[Any, bool]:
    """打开 BIFF 工作簿，返回 (book, 是否带格式信息)。

    ``formatting_info=True`` 才能拿到合并区域与字体（表头启发式的输入）；
    个别文件的格式记录损坏时退一步用纯值模式——丢表头识别好过整份文件报废。
    """
    import xlrd

    try:
        return xlrd.open_workbook(str(src), formatting_info=True), True
    except NotImplementedError:
        pass  # 极老的 BIFF 版本不支持 formatting_info，走纯值模式
    except xlrd.biffh.XLRDError as exc:
        _raise_classified(src, exc)
    except PermissionError as exc:
        raise DocFactoryError("E01", f"文件被占用或无读取权限：{src.name}") from exc
    except Exception:
        # 格式记录损坏等：还有机会用纯值模式读出数据，先不下结论
        pass

    try:
        return xlrd.open_workbook(str(src)), False
    except xlrd.biffh.XLRDError as exc:
        _raise_classified(src, exc)
        raise  # _raise_classified 必抛；这行只为让类型检查器安心
    except PermissionError as exc:
        raise DocFactoryError("E01", f"文件被占用或无读取权限：{src.name}") from exc
    except Exception as exc:
        raise DocFactoryError(
            "E01", f"「{src.name}」不是有效的 .xls 文件或已损坏（{type(exc).__name__}: {exc}）"
        ) from exc


def _raise_classified(src: Path, exc: Exception) -> None:
    """把 xlrd 的报错分类成用户能行动的 E 码。"""
    msg = str(exc).lower()
    if "encrypt" in msg or "password" in msg or "protected" in msg:
        raise DocFactoryError("E02", f"「{src.name}」受密码保护，请先解除密码再导入") from exc
    if "xlsx" in msg or "zip" in msg:
        raise DocFactoryError(
            "E01", f"「{src.name}」实际是新格式 Office 文件，请把扩展名改回 .xlsx 后重新导入"
        ) from exc
    raise DocFactoryError(
        "E01", f"「{src.name}」不是有效的 .xls 文件或已损坏（{exc}）"
    ) from exc


def _build_workbook(book: Any, has_formatting: bool) -> Any:
    from openpyxl import Workbook
    from openpyxl.styles import Font

    bold = Font(bold=True)
    wb = Workbook()
    wb.remove(wb.active)  # 去掉默认 "Sheet"，完全按源文件的 sheet 顺序与名字重建

    for si in range(book.nsheets):
        sheet = book.sheet_by_index(si)
        ws = wb.create_sheet(title=sheet.name)

        for r in range(sheet.nrows):
            for c, cell in enumerate(sheet.row(r)):
                value = _convert_value(cell, book.datemode)
                if value is None:
                    continue
                target = ws.cell(row=r + 1, column=c + 1, value=value)
                if has_formatting and _is_bold(book, sheet, r, c):
                    target.font = bold

        # merged_cells 是 0 基半开区间；openpyxl 要 1 基闭区间
        for rlo, rhi, clo, chi in getattr(sheet, "merged_cells", None) or []:
            if rhi - rlo <= 1 and chi - clo <= 1:
                continue
            ws.merge_cells(start_row=rlo + 1, start_column=clo + 1, end_row=rhi, end_column=chi)

    return wb


def _convert_value(cell: Any, datemode: int) -> Any:
    """xlrd 单元格 → openpyxl 可写值；空值与空串返回 None（不写入，产物更瘦）。"""
    ctype = cell.ctype
    if ctype == _XL_TEXT:
        return cell.value or None
    if ctype == _XL_NUMBER:
        return cell.value
    if ctype == _XL_DATE:
        return _convert_date(cell.value, datemode)
    if ctype == _XL_BOOLEAN:
        return bool(cell.value)
    if ctype == _XL_ERROR:
        from xlrd.biffh import error_text_from_code

        return error_text_from_code.get(cell.value, "#ERR!")
    return None  # EMPTY / BLANK


def _convert_date(value: float, datemode: int) -> Any:
    """Excel 序列日期 → date/time/datetime，让 xlsx_parser._fmt 输出干净的 ISO 文本。

    序列值 < 1 是纯时间；整日给 date（避免下游多出一截 " 00:00:00"）；其余 datetime。
    1900 闰年 bug 等越界值转不动时，退回原始数值——数字总比丢内容强。
    """
    from xlrd.xldate import xldate_as_datetime

    try:
        dt = xldate_as_datetime(value, datemode)
    except Exception:
        return value
    if 0 <= value < 1:
        return dt.time()
    if dt.time() == time.min:
        return dt.date()
    return dt


def _is_bold(book: Any, sheet: Any, row: int, col: int) -> bool:
    """加粗判定（尽力而为）：xf → font 两级索引里任何一环缺失都按不加粗算。"""
    try:
        xf = book.xf_list[sheet.cell_xf_index(row, col)]
        font = book.font_list[xf.font_index]
        return bool(font.bold) or int(font.weight or 0) >= 700
    except Exception:
        return False
