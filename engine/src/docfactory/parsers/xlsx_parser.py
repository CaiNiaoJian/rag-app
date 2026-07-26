"""xlsx 解析器（openpyxl，MIT）—— 严格实现 03 章 §7 的 Excel 契约表。

| 契约项 | 本文件的落实 |
|---|---|
| 公式单元格 | ``data_only=True`` 取缓存值；无缓存值时回退公式原文并记 warning |
| 合并单元格 | IR cell 带 rowspan/colspan，取值归左上格 |
| 数据区域 | 每 sheet 按连续非空区域切成 ``sheet_region``；表头行启发式识别 |
| 超大表 | 单 sheet > 10 万行截断并记 E05 变体 warning |
| 内嵌图片/图表 | 图片落 assets；图表只提标题与源数据区域引用，不还原图形 |
| 多 sheet | 全部解析，空 sheet 跳过 |

**页码**：Excel 的分页取决于打印设置，文件里没有稳定的页概念，因此 ``prov.page`` 用
**sheet 序号**占位（第 1 个 sheet = 1）。这样预览/溯源能定位到工作表，
又不会捏造一个用户在 Excel 里看不到的页号。

**为什么可能加载两遍工作簿**：``data_only=True`` 时公式单元格若从未被 Excel 计算保存过，
openpyxl 只能给出 None —— 此时无法区分「空格」与「没算过的公式」。所以只在**首次**
于数据区域内遇到空洞时，才惰性再开一次 ``data_only=False`` 的工作簿去问「这里是不是公式」。
绝大多数文件不触发；触发也只多一次加载。超大文件（> 64MB）直接放弃这条回退，
保住 03 章 §8 的内存预算。
"""

from __future__ import annotations

from contextlib import suppress
from datetime import date, datetime, time, timedelta
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from docfactory.errors import DocFactoryError
from docfactory.ir import IRNode, NodeContent, Prov, TableCell, TableContent
from docfactory.parsers import ParseEnv, assert_ooxml_readable

# 03 章 §7：单 sheet 超过 10 万行截断
MAX_SHEET_ROWS = 100_000
# 列上限：openpyxl 的 max_column 会被「整列设过格式」撑到 16384，扫过去纯属浪费
MAX_SHEET_COLS = 1024
# 惰性加载公式工作簿的体积闸门（见模块 docstring）
FORMULA_FALLBACK_MAX_BYTES = 64 * 1024 * 1024


def _open(src: Path, *, data_only: bool) -> Any:
    assert_ooxml_readable(src)
    try:
        return load_workbook(str(src), data_only=data_only, keep_links=False)
    except DocFactoryError:
        raise
    except Exception as exc:
        name = type(exc).__name__
        if "Encrypt" in name or "password" in str(exc).lower():
            raise DocFactoryError("E02", f"「{src.name}」受密码保护，请先解除密码再导入") from exc
        raise DocFactoryError("E01", f"打开 xlsx 失败：{src.name}（{name}: {exc}）") from exc


# ---------------------------------------------------------------- 值格式化


def _fmt(value: Any) -> str:
    """单元格值 → 文本。数值去掉浮点噪声，日期统一 ISO，方便下游切片与 LLM 消费。"""
    if value is None:
        return ""
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    if isinstance(value, datetime):
        return value.isoformat(sep=" ", timespec="seconds")
    if isinstance(value, (date, time)):
        return value.isoformat()
    if isinstance(value, timedelta):
        return str(value)
    if isinstance(value, float):
        if value.is_integer():
            return str(int(value))
        return f"{round(value, 10):g}"
    return str(value).strip()


def _is_blank(value: Any) -> bool:
    return value is None or (isinstance(value, str) and not value.strip())


# ---------------------------------------------------------------- 区域切分


class _Region:
    __slots__ = ("r0", "r1", "c0", "c1")

    def __init__(self, r0: int, r1: int, c0: int, c1: int):
        self.r0, self.r1, self.c0, self.c1 = r0, r1, c0, c1

    @property
    def ref(self) -> str:
        return (
            f"{get_column_letter(self.c0)}{self.r0}:{get_column_letter(self.c1)}{self.r1}"
        )


def _scan_regions(ws: Any, max_row: int, max_col: int) -> list[_Region]:
    """把 sheet 切成若干连续非空矩形区域（03 章 §7「数据区域」）。

    规则：整行为空 → 纵向断开；组内整列为空 → 横向断开。
    只记录行边界与列占用集合，不缓存单元格，10 万行也就几 MB。
    """
    regions: list[_Region] = []
    start: int | None = None
    end = 0
    used: set[int] = set()

    def close() -> None:
        nonlocal start, used
        if start is not None and used:
            for c0, c1 in _split_columns(used):
                regions.append(_Region(start, end, c0, c1))
        start, used = None, set()

    for row in ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
        row_cols = [cell.column for cell in row if not _is_blank(cell.value)]
        if not row_cols:
            close()
            continue
        if start is None:
            start = row[0].row
        end = row[0].row
        used.update(row_cols)
    close()
    return regions


def _split_columns(used: set[int]) -> list[tuple[int, int]]:
    """占用列集合 → 连续列段（中间隔着空列就断开）。"""
    out: list[tuple[int, int]] = []
    cols = sorted(used)
    c0 = prev = cols[0]
    for c in cols[1:]:
        if c > prev + 1:
            out.append((c0, prev))
            c0 = c
        prev = c
    out.append((c0, prev))
    return out


# ---------------------------------------------------------------- 表头启发式


def _header_row(ws: Any, region: _Region,
                merges: dict[tuple[int, int], tuple[int, int, bool]]) -> int | None:
    """返回表头所在的行号（绝对行），没有表头则 None（03 章 §7 表头行启发式）。

    候选不只是区域首行：真实报表极常见「首行是跨整个区域的合并标题（"季度报表"），
    第二行才是字段名」。所以先跳过这种横贯合并的标题行，再判定。
    """
    for row in (region.r0, region.r0 + 1):
        if row + 1 > region.r1:
            break
        if row == region.r0 and _is_banner_row(region, row, merges):
            continue      # 跨整个区域宽度的合并行 → 是标题不是表头
        if _looks_like_header(ws, region, row):
            return row
    return None


def _is_banner_row(region: _Region, row: int,
                   merges: dict[tuple[int, int], tuple[int, int, bool]]) -> bool:
    span = merges.get((row, region.c0))
    return bool(span and span[2] and span[1] >= region.c1 - region.c0 + 1)


def _looks_like_header(ws: Any, region: _Region, row: int) -> bool:
    """该行是否像表头：文本行 + 次行类型突变，或该行加粗而次行不加粗。"""
    if row + 1 > region.r1:
        return False
    first = [ws.cell(row, c) for c in range(region.c0, region.c1 + 1)]
    second = [ws.cell(row + 1, c) for c in range(region.c0, region.c1 + 1)]

    first_vals = [c.value for c in first if not _is_blank(c.value)]
    second_vals = [c.value for c in second if not _is_blank(c.value)]
    if not first_vals:
        return False

    all_text = all(isinstance(v, str) for v in first_vals)
    if all_text and second_vals and any(not isinstance(v, str) for v in second_vals):
        return True  # 类型突变：文本表头 + 数值/日期数据

    bold_first = sum(1 for c in first if _bold(c))
    bold_second = sum(1 for c in second if _bold(c))
    return all_text and bold_first >= max(1, len(first_vals) // 2) and bold_second == 0


def _bold(cell: Any) -> bool:
    try:
        return bool(cell.font and cell.font.bold)
    except Exception:
        return False


# ---------------------------------------------------------------- 合并单元格


def _merge_index(ws: Any) -> dict[tuple[int, int], tuple[int, int, bool]]:
    """(row, col) → (rowspan, colspan, 是否合并区左上格)。非左上格的 span 无意义。"""
    index: dict[tuple[int, int], tuple[int, int, bool]] = {}
    try:
        ranges = list(ws.merged_cells.ranges)
    except Exception:
        return index
    for rng in ranges:
        rowspan = rng.max_row - rng.min_row + 1
        colspan = rng.max_col - rng.min_col + 1
        for r in range(rng.min_row, rng.max_row + 1):
            for c in range(rng.min_col, rng.max_col + 1):
                origin = r == rng.min_row and c == rng.min_col
                index[(r, c)] = (rowspan, colspan, origin)
    return index


# ---------------------------------------------------------------- 主流程


class _FormulaProbe:
    """公式原文回退：惰性打开第二份工作簿，只在数据区出现空洞时才付这笔成本。"""

    def __init__(self, src: Path, env: ParseEnv):
        self.src = src
        self.env = env
        self._wb: Any = None
        self._tried = False
        self._reported = False

    def close(self) -> None:
        """及时释放第二份工作簿：它可能占到源文件几倍的内存，不该活到 GC 高兴为止。"""
        wb, self._wb = self._wb, None
        if wb is not None:
            with suppress(Exception):
                wb.close()

    def formula_at(self, sheet_title: str, row: int, col: int) -> str:
        if self._wb is None:
            if self._tried:
                return ""
            self._tried = True
            try:
                if self.src.stat().st_size > FORMULA_FALLBACK_MAX_BYTES:
                    return ""
                self._wb = load_workbook(str(self.src), data_only=False, keep_links=False)
            except Exception:
                return ""
        try:
            value = self._wb[sheet_title].cell(row, col).value
        except Exception:
            return ""
        if isinstance(value, str) and value.startswith("="):
            if not self._reported:
                self._reported = True
                self.env.note_warning(
                    "E05",
                    f"工作表「{sheet_title}」存在未计算的公式单元格，已回退显示公式原文；"
                    f"用 Excel 打开并保存一次即可得到计算结果",
                )
            return value
        return ""


def _emit_region(
    ws: Any, region: _Region, sheet_node: IRNode, page: int,
    env: ParseEnv, merges: dict[tuple[int, int], tuple[int, int, bool]],
    probe: _FormulaProbe,
) -> None:
    header_row = _header_row(ws, region, merges)
    cells: list[TableCell] = []
    for row in ws.iter_rows(
        min_row=region.r0, max_row=region.r1, min_col=region.c0, max_col=region.c1
    ):
        for cell in row:
            key = (cell.row, cell.column)
            rowspan, colspan, origin = merges.get(key, (1, 1, True))
            if not origin:
                continue  # 合并区的非左上格不产出（ir.table_to_grid 会补空）
            text = _fmt(cell.value)
            if not text and cell.value is None:
                text = probe.formula_at(ws.title, cell.row, cell.column)
            if not text and (rowspan, colspan) == (1, 1):
                continue  # 真空格不占 cell，减小 IR 体积
            cells.append(TableCell(
                r=cell.row - region.r0,
                c=cell.column - region.c0,
                rowspan=min(rowspan, region.r1 - cell.row + 1),
                colspan=min(colspan, region.c1 - cell.column + 1),
                text=text,
                is_header=header_row is not None and cell.row == header_row,
            ))
    if not cells:
        return
    env.builder.add(
        "sheet_region",
        parent=sheet_node,
        content=NodeContent(table=TableContent(cells=cells), range=region.ref),
        prov=[Prov(page=page)],
    )


def _emit_images_and_charts(ws: Any, sheet_node: IRNode, page: int, env: ParseEnv) -> None:
    """内嵌图片落 assets；图表只记标题与源数据引用（03 章 §7）。

    走的是 openpyxl 私有属性（``_images``/``_charts``，公开 API 没有暴露），
    因此整段包在 try 里：拿不到图就少几个 figure 节点，绝不能连正文一起赔进去。
    """
    try:
        images = list(getattr(ws, "_images", []) or [])
    except Exception:
        images = []
    for image in images:
        try:
            blob = image._data()
            ext = f".{(getattr(image, 'format', None) or 'png')}".lower()
        except Exception:
            continue
        if blob:
            ref = env.save_asset(blob, ext)
            env.builder.add(
                "figure", parent=sheet_node,
                content=NodeContent(image_ref=ref, ocr=False), prov=[Prov(page=page)],
            )

    try:
        charts = list(getattr(ws, "_charts", []) or [])
    except Exception:
        charts = []
    for chart in charts:
        title = _chart_title(chart)
        refs = _chart_refs(chart)
        text = f"[图表] {title or '(无标题)'}" + (f"　数据源：{', '.join(refs)}" if refs else "")
        env.builder.add(
            "paragraph", parent=sheet_node, content=NodeContent(text=text), prov=[Prov(page=page)]
        )


def _chart_title(chart: Any) -> str:
    try:
        title = chart.title
        if title is None:
            return ""
        texts: list[str] = []
        for para in title.tx.rich.p:
            for run in para.r or []:
                if run.t:
                    texts.append(run.t)
        return "".join(texts).strip()
    except Exception:
        return ""


def _chart_refs(chart: Any) -> list[str]:
    refs: list[str] = []
    try:
        for ser in chart.series:
            for holder in (getattr(ser, "val", None), getattr(ser, "cat", None)):
                ref = getattr(getattr(holder, "numRef", None), "f", None)
                if ref:
                    refs.append(str(ref))
    except Exception:
        return refs
    return refs


def parse(src: Path, env: ParseEnv) -> None:
    wb = _open(src, data_only=True)
    probe = _FormulaProbe(Path(src), env)
    try:
        _parse_sheets(wb, probe, env)
    finally:
        # 取消/异常（异常还会触发上层的 L2 兜底，届时又要开一份工作簿）时同样要放手，
        # 否则两三份工作簿同时驻留内存，直接顶穿 03 章 §8 的内存预算。
        probe.close()
        with suppress(Exception):
            wb.close()


def _parse_sheets(wb: Any, probe: _FormulaProbe, env: ParseEnv) -> None:
    titles = list(wb.sheetnames)
    total = max(1, len(titles))
    env.page_count = len(titles)

    for page, title in enumerate(titles, start=1):
        ws = wb[title]
        max_row = min(int(ws.max_row or 0), MAX_SHEET_ROWS)
        max_col = min(int(ws.max_column or 0), MAX_SHEET_COLS)
        if max_row <= 0 or max_col <= 0:
            env.progress(page, total)
            continue

        if (ws.max_row or 0) > MAX_SHEET_ROWS:
            env.note_warning(
                "E05",
                f"工作表「{title}」共 {ws.max_row} 行，已截断至 {MAX_SHEET_ROWS} 行，建议拆分文件",
                page=page,
            )
        if (ws.max_column or 0) > MAX_SHEET_COLS:
            env.note_warning(
                "E05",
                f"工作表「{title}」列数超过 {MAX_SHEET_COLS}，已截断",
                page=page,
            )

        regions = _scan_regions(ws, max_row, max_col)
        has_drawing = bool(getattr(ws, "_images", None)) or bool(getattr(ws, "_charts", None))
        if not regions and not has_drawing:
            env.progress(page, total)
            continue  # 空 sheet 跳过（03 章 §7）

        sheet_node = env.builder.add(
            "sheet", content=NodeContent(name=title), prov=[Prov(page=page)]
        )
        merges = _merge_index(ws)
        for region in regions:
            _emit_region(ws, region, sheet_node, page, env, merges, probe)
            env.check_cancel()
        _emit_images_and_charts(ws, sheet_node, page, env)

        env.mark_level(page, "L0")
        env.progress(page, total)
        env.check_cancel()


def parse_text_fallback(src: Path, env: ParseEnv) -> None:
    """L2 兜底：逐 sheet 把非空单元格按行拼成纯文本，放弃区域与合并信息。"""
    wb = _open(src, data_only=True)
    try:
        _fallback_sheets(wb, env)
    finally:
        with suppress(Exception):
            wb.close()


def _fallback_sheets(wb: Any, env: ParseEnv) -> None:
    env.page_count = len(wb.sheetnames)
    for page, title in enumerate(wb.sheetnames, start=1):
        ws = wb[title]
        max_row = min(int(ws.max_row or 0), MAX_SHEET_ROWS)
        max_col = min(int(ws.max_column or 0), MAX_SHEET_COLS)
        lines: list[str] = []
        if max_row > 0 and max_col > 0:
            for row in ws.iter_rows(
                min_row=1, max_row=max_row, min_col=1, max_col=max_col, values_only=True
            ):
                line = "\t".join(_fmt(v) for v in row).strip()
                if line:
                    lines.append(line)
        if not lines:
            continue
        node = env.builder.add(
            "sheet", content=NodeContent(name=title), prov=[Prov(page=page)]
        )
        env.builder.add(
            "paragraph", parent=node,
            content=NodeContent(text="\n".join(lines)), prov=[Prov(page=page)],
        )
        env.mark_level(page, "L2")
