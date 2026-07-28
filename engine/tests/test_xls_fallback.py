"""`.xls` xlrd 兜底路径（parsers/xls_compat.py + office_convert 接入）。

覆盖点：无 LibreOffice 时 .xls 仍有完整产出（值/合并/表头/日期）、
convert_chain 溯源标记、密码与损坏文件的分类报错、.doc/.ppt 不受影响仍报 E03。

fixture 用 xlwt（BSD，dev 依赖）程序化生成——与 corpus 的纪律一致：
标注即生成脚本的意图，样本与断言永远不会漂移。
"""

from __future__ import annotations

from datetime import date
from pathlib import Path

import pytest

from docfactory.config import Settings
from docfactory.errors import DocFactoryError
from docfactory.ir import table_to_grid
from docfactory.parsers import parse_document
from docfactory.parsers.xls_compat import convert_xls_to_xlsx


@pytest.fixture(autouse=True)
def no_soffice(monkeypatch: pytest.MonkeyPatch) -> None:
    """本文件全部用例都在「干净 Windows 机」的假设下跑：找不到 soffice。"""
    monkeypatch.setattr("docfactory.parsers.office_convert.find_soffice", lambda: None)


def _make_budget_xls(path: Path) -> None:
    """合并标题 + 加粗表头 + 数值/日期/布尔列的典型报表。"""
    import xlwt

    wb = xlwt.Workbook()
    ws = wb.add_sheet("预算")
    bold = xlwt.easyxf("font: bold on")
    ymd = xlwt.easyxf(num_format_str="yyyy-mm-dd")

    ws.write_merge(0, 0, 0, 3, "部门预算表")  # 横贯合并标题（banner 行）
    for col, text in enumerate(("科目", "预算", "签约日", "已审批")):
        ws.write(1, col, text, bold)
    rows = [
        ("人力", 1200000, date(2026, 2, 11), True),
        ("市场", 500000, date(2026, 3, 5), False),
        ("研发", 2000000, date(2026, 3, 28), True),
    ]
    for r, (name, budget, day, ok) in enumerate(rows, start=2):
        ws.write(r, 0, name)
        ws.write(r, 1, budget)
        ws.write(r, 2, day, ymd)
        ws.write(r, 3, ok)

    detail = wb.add_sheet("明细")
    detail.write(0, 0, "备注")
    detail.write(1, 0, "以上金额均为含税价")
    wb.save(str(path))


def test_parse_xls_without_soffice(paths, tmp_path: Path) -> None:
    """端到端：无 soffice 的 .xls 走 xlrd 兜底，产出与 xlsx 同构的 IR。"""
    src = tmp_path / "budget.xls"
    _make_budget_xls(src)

    ir = parse_document(
        src=src, doc_id="doc-xls", fmt="xls", paths=paths, settings=Settings(), ctx=None
    )

    assert ir.doc.convert_chain == ["xls->xlsx(xlrd)"]
    types = {n.type for n in ir.nodes}
    assert "sheet" in types and "sheet_region" in types

    regions = [n for n in ir.nodes if n.type == "sheet_region" and n.content.table]
    assert regions, "应产出至少一个数据区域"
    grid = table_to_grid(regions[0].content.table)
    flat = "\n".join(cell for row in grid for cell in row)
    assert "部门预算表" in flat            # 合并标题保留
    assert "研发" in flat and "2000000" in flat
    assert "2026-02-11" in flat            # 日期以 ISO 文本呈现
    assert "TRUE" in flat                  # 布尔值

    # 加粗表头 + 次行类型突变 → 表头启发式应命中「科目」行
    header_cells = [c for c in regions[0].content.table.cells if c.is_header]
    assert any(c.text == "科目" for c in header_cells), "加粗信息应穿透转换供表头启发式使用"


def test_merged_cells_survive_conversion(tmp_path: Path) -> None:
    """转换器单独验证：合并区域坐标换算正确（0 基半开 → 1 基闭）。"""
    from openpyxl import load_workbook

    src = tmp_path / "merged.xls"
    _make_budget_xls(src)
    produced = convert_xls_to_xlsx(src, tmp_path / "out")

    wb = load_workbook(str(produced))
    merged = {str(r) for r in wb["预算"].merged_cells.ranges}
    assert "A1:D1" in merged
    assert wb["预算"]["A1"].value == "部门预算表"
    assert wb["明细"]["A2"].value == "以上金额均为含税价"


def test_corrupt_ole_reports_e01(paths, tmp_path: Path) -> None:
    """OLE 头 + 垃圾字节：分类为 E01，不允许未分类异常裸抛。"""
    src = tmp_path / "broken.xls"
    src.write_bytes(b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1" + b"\x00" * 512)

    with pytest.raises(DocFactoryError) as exc_info:
        parse_document(
            src=src, doc_id="doc-bad", fmt="xls", paths=paths, settings=Settings(), ctx=None
        )
    assert exc_info.value.code == "E01"


def test_renamed_xlsx_gets_actionable_message(tmp_path: Path) -> None:
    """把 xlsx 改名成 .xls 是高频误操作：报错必须指路，而不是一句「损坏」。"""
    from openpyxl import Workbook

    fake = tmp_path / "actually_new_format.xls"
    wb = Workbook()
    wb.active["A1"] = "x"
    wb.save(str(fake))

    with pytest.raises(DocFactoryError) as exc_info:
        convert_xls_to_xlsx(fake, tmp_path / "out")
    assert exc_info.value.code == "E01"
    assert ".xlsx" in exc_info.value.detail


def test_doc_without_soffice_still_e03(paths, tmp_path: Path) -> None:
    """兜底只覆盖 .xls：.doc 在无 soffice 时仍按原样报 E03（消息含另存建议）。"""
    src = tmp_path / "legacy.doc"
    src.write_bytes(b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1" + b"\x00" * 128)

    with pytest.raises(DocFactoryError) as exc_info:
        parse_document(
            src=src, doc_id="doc-doc", fmt="doc", paths=paths, settings=Settings(), ctx=None
        )
    assert exc_info.value.code == "E03"
    assert "docx" in exc_info.value.detail  # 「另存为 .docx」的可操作指引仍在
